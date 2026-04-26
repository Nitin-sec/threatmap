"""
cli_menu.py — ThreatMap Infra Post-Scan Interactive Menu

Provides:
  - PostScanMenu: arrow-key navigation using questionary
  - open_in_browser(): opens HTML in best available browser
  - export_libreoffice(): converts TXT → ODT if LibreOffice available,
                          otherwise copies TXT with instructions
  - show_logs(): prints path and first N lines of scan log

Designed to loop until user selects Exit.
Never crashes — all actions wrapped in try/except.
"""

import os
import shutil
import subprocess
import sys
import json
from pathlib import Path
from typing import Optional

import questionary
from rich.console import Console
from rich.panel import Panel

from core.scan_logger import get_logger

log     = get_logger("menu")
console = Console()

Q = questionary.Style([
    ("qmark",       "fg:red bold"),
    ("question",    "fg:white bold"),
    ("answer",      "fg:cyan bold"),
    ("pointer",     "fg:red bold"),
    ("highlighted", "fg:cyan bold"),
    ("selected",    "fg:cyan"),
])

# ── Helpers ───────────────────────────────────────────────────────────────────

def _ok(msg):   console.print(f"  [bold green][[+]][/bold green]  {msg}")
def _w(msg):    console.print(f"  [bold yellow][[!]][/bold yellow]  {msg}")
def _e(msg):    console.print(f"  [bold red][[-]][/bold red]  {msg}")
def _i(msg):    console.print(f"  [bold blue][[*]][/bold blue]  {msg}")


def open_in_browser(html_path: str) -> bool:
    """Open HTML in best available browser. Returns True if opened."""
    if not html_path or not Path(html_path).is_file():
        _e(f"HTML report not found: {html_path}")
        return False

    browsers = ["firefox","chromium","chromium-browser","google-chrome",
                "brave-browser","opera","xdg-open"]
    for browser in browsers:
        if shutil.which(browser):
            try:
                subprocess.Popen(
                    [browser, html_path],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                _ok(f"Opened in [bold]{browser}[/bold]")
                _i(f"Path: [cyan]{html_path}[/cyan]")
                return True
            except Exception as exc:
                log.warning("[menu] browser %s failed: %s", browser, exc)
                continue

    _w("No browser found. Open this file manually:")
    console.print(f"  [cyan]{html_path}[/cyan]")
    return False


def _open_file(path: str) -> bool:
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
            return True
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        if shutil.which(opener):
            subprocess.Popen([opener, path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return True
    except Exception as exc:
        log.warning("[menu] file open failed: %s", exc)
    return False


def export_excel(json_path: str, output_dir: str) -> Optional[str]:
    """
    Export vulnerability data to Excel (.xlsx) from JSON report output.
    """
    if not json_path or not Path(json_path).is_file():
        _e(f"JSON report not found: {json_path}")
        return None

    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    except Exception:
        _e("Excel export unavailable: install/openpyxl dependency issue.")
        return None

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    findings = report_data.get("findings", [])
    summary = report_data.get("summary", {})
    meta = report_data.get("meta", {})
    total_findings = int(report_data.get("total", len(findings)))
    total_hosts = len({f"{f.get('host', '')}" for f in findings if f.get("host")})

    workbook = openpyxl.Workbook()

    # Sheet 1: Executive Summary
    ws_summary = workbook.active
    ws_summary.title = "Executive Summary"

    # Title
    ws_summary['A1'] = "THREATMAP"
    ws_summary['A1'].font = Font(size=24, bold=True)
    ws_summary.merge_cells('A1:E1')

    # Meta info
    ws_summary['A3'] = "Target:"
    ws_summary['B3'] = meta.get("target", "Unknown")
    ws_summary['A4'] = "Date:"
    ws_summary['B4'] = meta.get("generated_at", "")[:10]
    ws_summary['A5'] = "Scan Mode:"
    ws_summary['B5'] = meta.get("scan_mode", "").title()

    # Severity counts with colors
    ws_summary['A7'] = "Severity"
    ws_summary['B7'] = "Count"
    ws_summary['A7'].font = Font(bold=True)
    ws_summary['B7'].font = Font(bold=True)

    severities = [("Critical", "FFC7CE"), ("High", "FFDAB9"), ("Medium", "FFFF99"), ("Low", "C6EFCE"), ("Info", "BDD7EE")]
    row = 8
    for sev, color in severities:
        count = int(summary.get(sev, 0))
        ws_summary[f'A{row}'] = sev
        ws_summary[f'B{row}'] = count
        ws_summary[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    # Auto width for summary
    for col in ['A', 'B']:
        ws_summary.column_dimensions[col].width = 15

    # Sheet 2: Findings
    ws = workbook.create_sheet("Findings")
    headers = ["No", "Finding", "Host", "Details", "Severity", "Remediation"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.border = Border(bottom=Side(style='thin'))

    severity_colors = {
        "Critical": "FFC7CE",
        "High": "FFDAB9",
        "Medium": "FFFF99",
        "Low": "C6EFCE",
        "Info": "BDD7EE"
    }

    for i, finding in enumerate(findings, 1):
        row = [
            i,
            finding.get("observation", ""),
            f"{finding.get('host', '')}:{finding.get('port', '')}",
            finding.get("detail", ""),
            finding.get("severity", ""),
            finding.get("remediation", "")
        ]
        ws.append(row)
        # Color the severity cell
        sev_cell = ws.cell(row=i+1, column=5)
        color = severity_colors.get(finding.get("severity", ""), "FFFFFF")
        sev_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Auto column width
    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 12), 48)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    xlsx_path = out_dir / f"{Path(json_path).stem}.xlsx"
    workbook.save(xlsx_path)
    _ok(f"Excel report saved: [cyan]{xlsx_path}[/cyan]")
    if not _open_file(str(xlsx_path)):
        _i(f"Open manually: [cyan]{xlsx_path}[/cyan]")
    return str(xlsx_path)


def show_logs(log_path: str, tail_lines: int = 30) -> None:
    """Print log file path and tail of scan log."""
    if not log_path or not Path(log_path).is_file():
        _w(f"Log file not found: {log_path}")
        return

    _ok(f"Log file: [cyan]{log_path}[/cyan]")
    console.print()

    try:
        lines = Path(log_path).read_text(encoding="utf-8", errors="replace").splitlines()
        clean = []
        for line in lines:
            if " DEBUG " in line:
                continue
            lowered = line.lower()
            if any(x in lowered for x in [" cmd:", "exit code", "traceback", "threatmap.triage  [slm]", "provider:"]):
                continue
            clean.append(line)
        tail  = clean[-tail_lines:] if len(clean) > tail_lines else clean

        if len(clean) > tail_lines:
            _i(f"Showing last {tail_lines} of {len(clean)} log lines")
        console.print()
        for line in tail:
            # Colour-code log levels for readability
            if " ERROR " in line or " CRITICAL " in line:
                console.print(f"  [red]{line}[/red]")
            elif " WARNING " in line:
                console.print(f"  [yellow]{line}[/yellow]")
            elif " INFO " in line:
                console.print(f"  [dim]{line}[/dim]")
            else:
                console.print(f"  [dim]{line}[/dim]")
        console.print()
    except Exception as exc:
        _e(f"Could not read log: {exc}")


def show_report_paths(paths: dict[str, str]) -> None:
    """Print all generated report paths in a clean table."""
    console.print(Panel.fit("[bold green]Generated Reports[/bold green]", expand=False))
    labels = {"html":"HTML Report ","txt":"Text Report ","json":"JSON Data   "}
    for fmt, path in paths.items():
        label = labels.get(fmt, fmt.upper())
        exists = Path(path).is_file() if path else False
        icon   = "[green]✔[/green]" if exists else "[red]✗[/red]"
        console.print(f"  {icon}  {label} → [cyan]{path}[/cyan]")
    console.print()


# ── Post-scan menu ────────────────────────────────────────────────────────────

class PostScanMenu:
    """
    Interactive post-scan menu. Loops until user selects Exit.

    Args:
        report_paths: dict with keys 'html', 'txt', 'json'
        log_path:     path to scan.log
        output_dir:   directory for any additional exports
    """

    def __init__(
        self,
        report_paths: dict[str, str],
        log_path: str,
        output_dir: str,
    ) -> None:
        self.paths      = report_paths
        self.log_path   = log_path
        self.output_dir = output_dir

    def run(self) -> str:
        """Enter the menu loop. Returns 'continue' or 'exit'."""
        show_report_paths(self.paths)

        # Auto-open HTML immediately on first entry
        html = self.paths.get("html","")
        if html and Path(html).is_file():
            _i("Opening HTML report in browser...")
            open_in_browser(html)
            console.print()

        console.print(Panel.fit("[dim]What would you like to do?[/dim]", expand=False))
        console.print()

        while True:
            try:
                choice = questionary.select(
                    "  Select action:",
                    choices=self._build_choices(),
                    style=Q,
                ).ask()
            except (KeyboardInterrupt, EOFError):
                console.print()
                _i("Exiting.")
                return "exit"

            if choice is None or choice == "exit":
                console.print()
                _ok("Session complete. Stay safe.")
                console.print()
                return "exit"

            if choice == "continue":
                return "continue"

            console.print()
            self._handle(choice)
            console.print()

    def _build_choices(self) -> list[dict]:
        choices = []

        html = self.paths.get("html","")
        if html and Path(html).is_file():
            choices.append({"name": "📄  View HTML Report (browser)", "value": "html"})

        js = self.paths.get("json","")
        if js and Path(js).is_file():
            choices.append({"name": "📊  Export as Excel (.xlsx)", "value": "xlsx"})

        choices.append({"name": "📋  Show Report File Paths", "value": "paths"})
        choices.append({"name": "📄  View Scan Logs",         "value": "logs"})
        choices.append({"name": "🔁  Continue Scanning",     "value": "continue"})
        choices.append({"name": "🚪  Exit",                   "value": "exit"})

        return choices

    def _handle(self, choice: str) -> None:
        try:
            if choice == "html":
                html = self.paths.get("html","")
                if html:
                    open_in_browser(html)
                else:
                    _e("HTML report was not generated.")

            elif choice == "xlsx":
                js = self.paths.get("json","")
                if js:
                    export_excel(js, self.output_dir)
                else:
                    _e("JSON report was not generated.")

            elif choice == "paths":
                show_report_paths(self.paths)

            elif choice == "logs":
                show_logs(self.log_path)

        except Exception as exc:
            _e(f"Action failed: {exc}")
            log.error("[menu] action %s failed: %s", choice, exc)
