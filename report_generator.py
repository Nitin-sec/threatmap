"""
report_generator.py — ThreatMap Infra Report Generator

Matches the reference professional pentest report format:

Sheet 1 — Cover
  Confidential header, scan metadata, risk summary matrix

Sheet 2 — Findings   (matches reference exactly)
  Risk summary table → then:
  S.No | Observation Name | Detailed Observations | Impacted Module |
  Severity | Risk/Impact | Recommendation

Sheet 3 — Evidence
  Raw tool output, DNS data, nuclei findings, screenshots

Additional "synthetic" findings (not from open ports) are derived from:
  - Missing security headers (from HTTP evidence)
  - WAF absence detection
  - DNSSEC status (from dig output)
  - Nuclei CVE findings (parsed separately)
"""

import glob
import json
import logging
import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("threatmap.report")

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
SEV_BG = {
    "Critical": "C0392B", "High": "E74C3C",
    "Medium":   "E67E22", "Low":  "F39C12",
    "Info":     "2980B9",
}
SEV_FG = {k: "FFFFFF" for k in SEV_BG}

DARK_NAVY  = "1A252F"
SLATE      = "2C3E50"
LIGHT_GRAY = "F8F9FA"
ALT_GRAY   = "EDF0F2"
BORDER_CLR = "CCD1D9"
MID_GRAY   = "7F8C8D"
HEADER_BG  = "2C3E50"

# ---------------------------------------------------------------------------
# Sanitiser — strips ANSI + illegal xlsx chars
# ---------------------------------------------------------------------------
_ANSI_RE    = re.compile(r"\x1b(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])")
_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def _clean(text, max_chars=32_000) -> str:
    if not text:
        return ""
    text = _ANSI_RE.sub("", str(text))
    text = _ILLEGAL_RE.sub("", text)
    return text[:max_chars]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="1A252F", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, name="Calibri", size=size, italic=italic)

def _border(color: str = BORDER_CLR) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _set(ws, row, col, value=None, bold=False, color="1A252F", size=10,
         italic=False, bg=None, halign="left", valign="top", wrap=True) -> None:
    c = ws.cell(row=row, column=col, value=_clean(value) if value is not None else None)
    c.font      = _font(bold=bold, color=color, size=size, italic=italic)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = _fill(bg)
    c.border = _border()

# ---------------------------------------------------------------------------
# Synthetic findings — derived from scan artefacts, not open ports
# ---------------------------------------------------------------------------

def _synthetic_findings(domain: str, reports_dir: str) -> list[dict]:
    """
    Build extra findings from evidence data: missing headers, WAF, DNSSEC.
    These match the style of the reference report exactly.
    """
    findings = []

    # ── Missing security headers ───────────────────────────────────────
    ev_files = glob.glob(f"{reports_dir}/evidence_{domain}*.json")
    if not ev_files:
        ev_files = glob.glob(f"{reports_dir}/evidence_*.json")
    if ev_files:
        try:
            ev      = json.loads(Path(ev_files[0]).read_text())
            missing = ev.get("missing_security_headers", [])
            if missing:
                findings.append({
                    "observation_name":    "Missing HTTP Security Headers",
                    "detailed_observation": (
                        f"Important security headers not detected: "
                        f"{', '.join(missing[:5])}."
                    ),
                    "impacted_module":  "Web Server",
                    "severity":         "Low",
                    "risk_impact":      (
                        "Absence of security headers increases exposure to clickjacking, "
                        "MIME sniffing, and cross-site scripting attacks."
                    ),
                    "recommendation":   (
                        "Add the following headers to your web server configuration: "
                        "X-Frame-Options: DENY, X-Content-Type-Options: nosniff, "
                        "Strict-Transport-Security: max-age=31536000; includeSubDomains, "
                        "Content-Security-Policy, Referrer-Policy."
                    ),
                    "ai_enhanced": False,
                })
        except Exception:
            pass

    # ── WAF absence ────────────────────────────────────────────────────
    waf_path = f"{reports_dir}/wafw00f_{domain}.txt"
    try:
        waf_text = Path(waf_path).read_text()
        has_waf  = any(k in waf_text.lower() for k in ["is behind", "detected"])
        if not has_waf:
            findings.append({
                "observation_name":    "No Web Application Firewall Detected",
                "detailed_observation": (
                    "Automated WAF detection scan did not identify any WAF or "
                    "CDN security layer protecting the application."
                ),
                "impacted_module":  "Application Layer",
                "severity":         "Low",
                "risk_impact":      (
                    "Without a WAF, the application is directly exposed to automated "
                    "attack tools, SQL injection attempts, and vulnerability scanners."
                ),
                "recommendation":   (
                    "Deploy a WAF (e.g., Cloudflare, AWS WAF, ModSecurity) or route "
                    "traffic through a CDN security layer."
                ),
                "ai_enhanced": False,
            })
    except FileNotFoundError:
        pass

    # ── DNSSEC check ───────────────────────────────────────────────────
    dig_path = f"{reports_dir}/dig_{domain}.txt"
    try:
        dig_text = Path(dig_path).read_text().lower()
        if "dnssec" not in dig_text and "rrsig" not in dig_text:
            findings.append({
                "observation_name":    "DNSSEC Not Configured",
                "detailed_observation": (
                    "DNSSEC signing was not detected for the domain. "
                    "DNS records are not cryptographically authenticated."
                ),
                "impacted_module":  "DNS Infrastructure",
                "severity":         "Low",
                "risk_impact":      (
                    "Without DNSSEC, the domain is vulnerable to DNS cache poisoning "
                    "attacks that redirect users to malicious servers."
                ),
                "recommendation":   (
                    "Enable DNSSEC at your domain registrar. "
                    "Generate a KSK and ZSK key pair and publish DS records to the parent zone."
                ),
                "ai_enhanced": False,
            })
    except FileNotFoundError:
        pass

    # ── Nuclei high/critical findings ─────────────────────────────────
    nuclei_path = f"{reports_dir}/nuclei_{domain}.txt"
    try:
        nuclei_lines = [l.strip() for l in
                        Path(nuclei_path).read_text().splitlines() if l.strip()]
        for line in nuclei_lines[:5]:   # max 5 nuclei findings as synthetic rows
            # Nuclei output format: [severity] [template-id] [url] [matcher]
            parts = line.split("]")
            sev_raw = parts[0].lstrip("[").strip().title() if parts else "Medium"
            sev     = sev_raw if sev_raw in SEV_BG else "Medium"
            findings.append({
                "observation_name":    f"CVE / Misconfiguration — {parts[1].strip(' [') if len(parts) > 1 else 'See detail'}",
                "detailed_observation": _clean(line),
                "impacted_module":      "Web Application",
                "severity":             sev,
                "risk_impact":          "Nuclei template-matched finding — verified vulnerability or misconfiguration.",
                "recommendation":       "Review the finding, apply vendor patch or configuration fix. Cross-reference CVE advisory.",
                "ai_enhanced": False,
            })
    except FileNotFoundError:
        pass

    return findings


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def _build_cover(wb: Workbook, meta: dict, sev_counts: dict) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines  = False
    ws.sheet_view.showRowColHeaders = False

    for col, w in zip("ABCDEFGH", [2, 4, 28, 2, 28, 2, 14, 10]):
        ws.column_dimensions[col].width = w

    # Confidential banner
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value     = "CONFIDENTIAL — FOR AUTHORIZED RECIPIENTS ONLY"
    c.fill      = _fill(DARK_NAVY)
    c.font      = Font(bold=True, color="E74C3C", size=10, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Title
    ws.merge_cells("B4:H5")
    t = ws["B4"]
    t.value     = "ThreatMap Infra — Vulnerability Assessment Report"
    t.font      = Font(bold=True, size=18, color=DARK_NAVY, name="Calibri")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 16

    ws.merge_cells("B6:H6")
    sub = ws["B6"]
    sub.value = (
        f"Target: {meta.get('target', '—')}   |   "
        f"Report Date: {datetime.now().strftime('%d %B %Y')}   |   "
        f"Mode: {(meta.get('scan_mode', 'balanced')).title()}"
    )
    sub.font      = Font(italic=True, size=10, color=MID_GRAY, name="Calibri")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B8:H8")
    ws["B8"].fill = _fill(DARK_NAVY)
    ws.row_dimensions[8].height = 3

    # Metadata
    ws.cell(row=10, column=3, value="Assessment Details").font = Font(
        bold=True, size=11, color=DARK_NAVY, name="Calibri"
    )
    rows_meta = [
        ("Target / Scope",    meta.get("target", "—")),
        ("Scan Mode",         meta.get("scan_mode", "balanced").title()),
        ("AI Triage Engine",  meta.get("ai_provider", "Rule-based")),
        ("Scan Started",      (meta.get("started_at", "") or "")[:19].replace("T", "  ")),
        ("Scan Completed",    (meta.get("completed_at") or "In Progress")[:19].replace("T", "  ")),
        ("Hosts Scanned",     str(meta.get("total_hosts", 0))),
        ("Total Findings",    str(meta.get("total_findings", 0))),
        ("Report Generated",  datetime.now().strftime("%d %B %Y, %H:%M")),
    ]
    for i, (label, value) in enumerate(rows_meta, start=11):
        ws.cell(row=i, column=3, value=label).font = Font(bold=True, size=10, color="444444", name="Calibri")
        ws.cell(row=i, column=5, value=str(value)).font = Font(size=10, color=DARK_NAVY, name="Calibri")
        ws.row_dimensions[i].height = 18

    # Risk summary
    ws.cell(row=10, column=7, value="Risk Summary").font = Font(
        bold=True, size=11, color=DARK_NAVY, name="Calibri"
    )
    for i, sev in enumerate(["Critical", "High", "Medium", "Low", "Info"], start=11):
        cnt = sev_counts.get(sev, 0)
        lc = ws.cell(row=i, column=7, value=sev)
        lc.fill = _fill(SEV_BG[sev])
        lc.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = _border()
        vc = ws.cell(row=i, column=8, value=cnt)
        vc.font = Font(bold=True, size=10, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = _border()

    # Methodology
    ws.cell(row=21, column=3, value="Methodology").font = Font(
        bold=True, size=11, color=DARK_NAVY, name="Calibri"
    )
    ws.merge_cells("C22:H25")
    sc = ws["C22"]
    sc.value = (
        "This automated vulnerability assessment used ThreatMap Infra to enumerate "
        "network-accessible services and web application security posture. "
        "Tools used: Nmap (port/service scanning), Nikto (web vulnerability scanning), "
        "Gobuster (directory enumeration), SSLScan (TLS analysis), cURL (HTTP headers), "
        "WhatWeb (technology fingerprinting), wafw00f (WAF detection), Nuclei (CVE and "
        "misconfiguration templates), dig/whois (DNS and registration). "
        "Findings are classified by CVSS v3.1 severity."
    )
    sc.font      = Font(size=10, color="333333", name="Calibri")
    sc.alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=27, column=3, value="Disclaimer").font = Font(
        bold=True, size=10, color="E74C3C", name="Calibri"
    )
    ws.merge_cells("C28:H29")
    dc = ws["C28"]
    dc.value = (
        "This report is generated by an automated scanning tool and should be validated "
        "by a qualified security professional before remediation decisions are made. "
        "Automated scanners may produce false positives and do not replace manual "
        "penetration testing. Unauthorized scanning is illegal."
    )
    dc.font      = Font(size=9, color=MID_GRAY, name="Calibri", italic=True)
    dc.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Sheet 2: Findings  — matches reference format exactly
# ---------------------------------------------------------------------------

def _build_findings(wb: Workbook, all_findings: list, sev_counts: dict) -> None:
    ws = wb.create_sheet("Findings")
    ws.sheet_view.showGridLines = False

    # ── Risk summary table (top) ───────────────────────────────────────
    # Header row
    for col, hdr in enumerate(["Risk Rating", "No. of Issues Identified"], start=1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font      = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
        c.fill      = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[1].height = 22

    # Severity rows
    for i, sev in enumerate(["Critical", "High", "Medium", "Low", "Info"], start=2):
        cnt = sev_counts.get(sev, 0)
        lc = ws.cell(row=i, column=1, value=sev)
        lc.fill      = _fill(SEV_BG[sev])
        lc.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = _border()
        vc = ws.cell(row=i, column=2, value=cnt)
        vc.font      = Font(size=10, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = _border()
        ws.row_dimensions[i].height = 18

    # Total row
    total_row = 7
    tc = ws.cell(row=total_row, column=1, value="Total Observations")
    tc.font = Font(bold=True, size=10, name="Calibri")
    tc.fill = _fill("EAECEE")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = _border()
    tv = ws.cell(row=total_row, column=2, value=sum(sev_counts.values()))
    tv.font = Font(bold=True, size=10, name="Calibri")
    tv.fill = _fill("EAECEE")
    tv.alignment = Alignment(horizontal="center", vertical="center")
    tv.border = _border()
    ws.row_dimensions[total_row].height = 20

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22

    # ── Findings table (row 10 onwards) ───────────────────────────────
    # Blank spacer rows
    hdr_row = 10

    headers    = ["S. No", "Observation Name", "Detailed Observations",
                  "Impacted Module", "Severity", "Risk / Impact", "Recommendation"]
    col_widths = [7, 32, 50, 18, 12, 45, 50]

    for col_i, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=hdr_row, column=col_i, value=hdr)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill      = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[hdr_row].height = 28
    ws.freeze_panes = f"A{hdr_row + 1}"

    # ── Finding rows ──────────────────────────────────────────────────
    for idx, f in enumerate(all_findings, start=1):
        r   = hdr_row + idx
        alt = idx % 2 == 0
        sev = str(f.get("severity") or "Info")
        bg  = ALT_GRAY if alt else "FFFFFF"

        values = [
            str(idx),
            _clean(f.get("observation_name") or "Security Finding"),
            _clean(f.get("detailed_observation") or f.get("risk_summary") or "—"),
            _clean(f.get("impacted_module") or "Network Service"),
            sev,
            _clean(f.get("risk_impact") or f.get("business_impact") or "—"),
            _clean(f.get("recommendation") or f.get("remediation") or "—"),
        ]

        for col_i, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.border    = _border()
            c.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if col_i in (1, 5) else "left"
            )
            if col_i == 5:   # Severity — coloured
                c.fill = _fill(SEV_BG.get(sev, "2980B9"))
                c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            elif col_i == 1:  # Row number
                c.font = Font(bold=True, size=10, name="Calibri")
                c.fill = _fill(bg)
            else:
                c.font = Font(size=10, name="Calibri")
                c.fill = _fill(bg)

        ws.row_dimensions[r].height = 70

    # ── Footer note ───────────────────────────────────────────────────
    note_row = hdr_row + len(all_findings) + 2
    ws.merge_cells(f"A{note_row}:G{note_row}")
    nc = ws.cell(row=note_row, column=1,
                 value="Note: Findings sourced from automated scanning tools. "
                       "Validate all findings before remediation. "
                       "AI-assisted findings use local SLM analysis — accuracy depends on model quality.")
    nc.font      = Font(size=8, color=MID_GRAY, name="Calibri", italic=True)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 24


# ---------------------------------------------------------------------------
# Sheet 3: Evidence
# ---------------------------------------------------------------------------

def _read_text(path: str, max_lines: int = 60) -> str:
    try:
        raw   = Path(path).read_text(encoding="utf-8", errors="replace")
        lines = [l.rstrip() for l in raw.splitlines() if l.strip()]
        return _clean("\n".join(lines[:max_lines]))
    except FileNotFoundError:
        return ""


def _build_evidence(wb: Workbook, host_rows: list, reports_dir: str) -> None:
    ws = wb.create_sheet("Evidence")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 90
    cur_row = 1

    def section(title: str) -> None:
        nonlocal cur_row
        ws.merge_cells(f"B{cur_row}:C{cur_row}")
        c = ws.cell(row=cur_row, column=2, value=_clean(title))
        c.fill = _fill(DARK_NAVY)
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

    def kv(label: str, value: str) -> None:
        nonlocal cur_row
        safe = _clean(str(value or ""))
        if not safe.strip():
            return
        ws.cell(row=cur_row, column=2, value=label).font = Font(
            bold=True, size=10, color="333333", name="Calibri")
        vc = ws.cell(row=cur_row, column=3, value=safe)
        vc.font = Font(size=10, name="Calibri")
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[cur_row].height = max(15, min(len(safe) // 6, 200))
        cur_row += 1

    def raw(label: str, content: str, color: str = DARK_NAVY) -> None:
        nonlocal cur_row
        safe = _clean(content)
        if not safe.strip():
            return
        ws.cell(row=cur_row, column=2, value=label).font = Font(
            bold=True, size=10, color="333333", name="Calibri")
        vc = ws.cell(row=cur_row, column=3, value=safe)
        vc.font = Font(size=9, name="Courier New", color=color)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        lines = safe.count("\n") + 1
        ws.row_dimensions[cur_row].height = max(15, min(lines * 13, 400))
        cur_row += 1

    def spacer(n: int = 1) -> None:
        nonlocal cur_row
        cur_row += n

    for host in host_rows:
        domain = host.get("domain", "")
        url    = host.get("url", "—")
        section(f"Host: {url}")

        # HTTP evidence
        ev_files = glob.glob(f"{reports_dir}/evidence_{domain}*.json")
        if not ev_files:
            ev_files = glob.glob(f"{reports_dir}/evidence_*.json")
        if ev_files:
            try:
                ev = json.loads(Path(ev_files[0]).read_text())
                kv("HTTP Status",    str(ev.get("status_code") or "—"))
                kv("Page Title",     ev.get("title") or "—")
                kv("Server",         ev.get("server") or "—")
                kv("Response Time",  f"{ev.get('response_time_ms', '—')} ms")
                missing = ev.get("missing_security_headers", [])
                if missing:
                    kv("Missing Security Headers",
                       "\n".join(f"  • {h}" for h in missing))
                if ev.get("redirect_chain"):
                    kv("Redirect Chain", " → ".join(ev["redirect_chain"]))
            except Exception:
                pass

        if host.get("ports_services"):
            kv("Open Ports / Services", host["ports_services"])

        raw("WHOIS",           _read_text(f"{reports_dir}/whois_{domain}.txt",    35))
        raw("DNS Records",     _read_text(f"{reports_dir}/dig_{domain}.txt",      40))

        # WAF — filter ASCII art
        waf_raw = _read_text(f"{reports_dir}/wafw00f_{domain}.txt", 20)
        if waf_raw:
            useful = [l for l in waf_raw.splitlines()
                      if any(k in l for k in
                             ["Checking", "behind", "detected", "No WAF",
                              "Generic", "identified", "[*]", "[+]", "[-]"])]
            raw("WAF Detection", "\n".join(useful) or waf_raw)

        raw("Nmap Port Scan",      _read_text(f"{reports_dir}/nmap_{domain}.xml",     70))
        raw("Nikto — Web Scan",    _read_text(f"{reports_dir}/nikto_{domain}.txt",    50))
        raw("Gobuster — Dirs",     _read_text(f"{reports_dir}/gobuster_{domain}.txt", 35))
        raw("SSLScan — TLS",       _read_text(f"{reports_dir}/sslscan_{domain}.txt",  45))
        raw("HTTP Headers",        _read_text(f"{reports_dir}/curl_headers_{domain}.txt", 25))

        # Nuclei — red highlight
        nuclei = _read_text(f"{reports_dir}/nuclei_{domain}.txt", 50)
        if nuclei:
            ws.cell(row=cur_row, column=2,
                    value="Nuclei — CVE & Misconfigs").font = Font(
                bold=True, size=10, color="C0392B", name="Calibri")
            vc = ws.cell(row=cur_row, column=3, value=nuclei)
            vc.font = Font(size=9, name="Courier New", color="C0392B")
            vc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[cur_row].height = max(15,
                min(nuclei.count("\n") * 13, 400))
            cur_row += 1

        spacer(2)

    # Screenshots
    pngs = []
    for rt, _, files in os.walk(reports_dir):
        for f in sorted(files):
            if f.lower().endswith(".png"):
                pngs.append(os.path.join(rt, f))

    if pngs:
        section("Visual Evidence — Screenshots")
        spacer(1)
        for png in pngs:
            try:
                img   = XLImage(png)
                max_w = 640
                if img.width > max_w:
                    r          = max_w / img.width
                    img.width  = int(img.width  * r)
                    img.height = int(img.height * r)
                ws.cell(row=cur_row, column=2,
                        value=os.path.basename(png)).font = Font(
                    bold=True, size=10, name="Calibri")
                cur_row += 1
                ws.add_image(img, f"B{cur_row}")
                rows_needed = max(20, img.height // 15 + 2)
                for rr in range(cur_row, cur_row + rows_needed):
                    ws.row_dimensions[rr].height = 15
                cur_row += rows_needed
                spacer(1)
            except Exception as exc:
                logger.warning("Screenshot embed failed: %s", exc)
    else:
        section("Visual Evidence — Screenshots")
        nc = ws.cell(row=cur_row + 1, column=2,
                     value="No screenshots. Install EyeWitness for visual evidence.")
        nc.font = Font(italic=True, size=10, color=MID_GRAY, name="Calibri")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_excel(
    db_path:     str = "threatmap.db",
    output:      str = "Findings.xlsx",
    scan_id:     int | None = None,
    reports_dir: str = "reports",
) -> str:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    if scan_id is None:
        row     = cur.execute("SELECT id FROM scans ORDER BY id DESC LIMIT 1").fetchone()
        scan_id = row["id"] if row else None

    meta: dict = {"ai_provider": "Rule-based"}
    if scan_id:
        sr = cur.execute(
            "SELECT target, scan_mode, started_at, completed_at FROM scans WHERE id=?",
            (scan_id,),
        ).fetchone()
        if sr:
            meta = dict(sr)
        meta["total_hosts"] = cur.execute(
            "SELECT COUNT(*) FROM hosts WHERE scan_id=?", (scan_id,)
        ).fetchone()[0]
        ai_row = cur.execute(
            "SELECT triage_method FROM triage WHERE ai_enhanced=1 LIMIT 1"
        ).fetchone()
        if ai_row:
            meta["ai_provider"] = f"Local SLM ({ai_row['triage_method']})"

    # Port-based findings from triage table
    triage_rows = cur.execute(
        """
        SELECT t.host, t.port, t.service, t.severity, t.cvss_score,
               t.risk_summary, t.remediation, t.attack_scenario,
               t.business_impact, t.false_positive_likelihood,
               t.actively_exploited, t.triage_method, t.priority_rank,
               t.ai_enhanced,
               -- new fields (may be NULL in older DBs)
               t.observation_name, t.detailed_observation,
               t.impacted_module, t.risk_impact
        FROM   triage t
        ORDER  BY t.priority_rank ASC, t.cvss_score DESC
        """
    ).fetchall()

    # Build finding dicts — prefer new fields, fall back to legacy
    port_findings = []
    for r in triage_rows:
        d = dict(r)
        # Fallback: if new fields not in DB yet, derive from legacy
        if not d.get("observation_name"):
            svc  = (d.get("service") or "").lower()
            port = str(d.get("port") or "")
            from ai_triage import RULE_OBSERVATIONS, DEFAULT_RULE
            rule = RULE_OBSERVATIONS.get(svc, DEFAULT_RULE)
            d["observation_name"]    = rule["observation_name"]
            d["impacted_module"]     = rule["impacted_module"]
            d["risk_impact"]         = d.get("business_impact") or rule["risk_impact"]
            d["detailed_observation"]= (
                d.get("risk_summary") or
                f"Port {port}/TCP ({svc.upper()}) is publicly accessible."
            )
        port_findings.append(d)

    # Synthetic findings from scan artefacts
    domain = meta.get("target", "")
    synthetic = _synthetic_findings(domain, reports_dir)

    all_findings = port_findings + synthetic
    meta["total_findings"] = len(all_findings)

    sev_counts: dict[str, int] = {}
    for f in all_findings:
        s = f.get("severity") or "Info"
        sev_counts[s] = sev_counts.get(s, 0) + 1

    # Host rows for evidence sheet
    host_rows = []
    if scan_id:
        rows = cur.execute(
            """
            SELECT h.url, h.domain,
                   GROUP_CONCAT(p.port || '/TCP (' || COALESCE(p.service,'?') || ')', '  •  ')
                       AS ports_services
            FROM   hosts h
            LEFT JOIN open_ports p ON h.id = p.host_id AND p.state = 'open'
            WHERE  h.scan_id = ?
            GROUP  BY h.id ORDER BY h.url
            """,
            (scan_id,),
        ).fetchall()
        host_rows = [dict(r) for r in rows]

    conn.close()

    wb = Workbook()
    _build_cover(wb, meta, sev_counts)
    _build_findings(wb, all_findings, sev_counts)
    _build_evidence(wb, host_rows, reports_dir)

    wb.save(output)
    abs_path = str(Path(output).resolve())
    logger.info("Report → %s  (%d findings)", abs_path, len(all_findings))
    return abs_path
